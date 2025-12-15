#!/usr/bin/env python3
"""
物料編碼申請管理系統 V3.5 優化版
Excel處理器 - 增強版

功能：
1. 將前端JSON資料轉換為標準SAP格式Excel
2. 支援八大類物料的不同欄位配置
3. 包裝說明欄位的完整處理
4. 資料驗證和錯誤處理
5. 批量處理支援

作者: System Development Team
版本: V3.5 Optimized
日期: 2024-11-21
"""

import pandas as pd
import json
import os
import sys
from datetime import datetime
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import logging
import argparse

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


class MaterialExcelProcessor:
    """優化版Excel處理器"""
    
    def __init__(self, config_path: Optional[str] = None):
        """
        初始化處理器
        
        Args:
            config_path: 配置檔案路徑（可選）
        """
        self.category_mapping = {
            'H': 'Handle',
            'S': 'Slide',
            'M': 'ModuleAssy',
            'D': 'DecorativeHardware',
            'F': 'FunctionalHardware',
            'B': 'BuildersHardware',
            'I': 'IndustrialPartsSolution',
            'O': 'Others'
        }
        
        # 定義各類別的欄位結構（優化版）
        self.category_columns = self._init_category_columns()
        
        # 載入配置
        if config_path and os.path.exists(config_path):
            self.config = self._load_config(config_path)
        else:
            self.config = self._get_default_config()
        
        # 初始化樣式
        self._init_styles()
        
        logger.info("Excel處理器初始化完成")
    
    def _init_category_columns(self) -> Dict[str, List[str]]:
        """初始化各類別的欄位結構"""
        base_columns = [
            '料號', '料件說明', '客戶說明', '產品大類', '產品中類', '產品小類',
            '料件基本材質', '料件外型長', '料件外型寬', '料件外型高', '料件外型重量'
        ]
        
        packaging_columns = [
            '個別產品包裝', '配件內容', '配件', '內盒', '外箱',
            '運輸與托盤要求', '裝櫃要求', 'Other'
        ]
        
        common_columns = [
            '料件表面處理', '料件顏色', '測試要求', '認證要求', '品質標準',
            'MOQ', '單位', '客戶參考號', '供應商編號', '建立日期', '狀態'
        ]
        
        return {
            'Handle': base_columns + ['把手長度', '孔距'] + common_columns + packaging_columns,
            'Slide': base_columns + ['滑軌長度', '滑軌載重', '滑軌類型', '鋼珠大小'] + common_columns + packaging_columns,
            'ModuleAssy': base_columns + ['模組類型', '適用櫃體寬度'] + common_columns + packaging_columns,
            'DecorativeHardware': base_columns + ['裝飾風格'] + common_columns + packaging_columns,
            'FunctionalHardware': base_columns + ['功能類型', '承載能力'] + common_columns + packaging_columns,
            'BuildersHardware': base_columns + ['建築應用', '安全等級'] + common_columns + packaging_columns,
            'IndustrialPartsSolution': base_columns + ['工業應用', '承載等級'] + common_columns + packaging_columns,
            'Others': base_columns + common_columns + packaging_columns
        }
    
    def _get_default_config(self) -> Dict:
        """取得預設配置"""
        return {
            'encoding': 'utf-8-sig',
            'date_format': '%Y-%m-%d',
            'decimal_places': 2,
            'max_rows_per_sheet': 65000,
            'include_summary': True,
            'include_validation': True,
            'auto_filter': True,
            'freeze_panes': 'B2'
        }
    
    def _load_config(self, config_path: str) -> Dict:
        """載入配置檔案"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.warning(f"載入配置失敗，使用預設配置: {e}")
            return self._get_default_config()
    
    def _init_styles(self):
        """初始化Excel樣式"""
        # 標題樣式
        self.header_style = NamedStyle(name='header')
        self.header_style.font = Font(bold=True, color='FFFFFF', size=11)
        self.header_style.fill = PatternFill(
            start_color='366092',
            end_color='366092',
            fill_type='solid'
        )
        self.header_style.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        self.header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='medium')
        )
        
        # 資料樣式
        self.data_style = NamedStyle(name='data')
        self.data_style.font = Font(size=10)
        self.data_style.alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True
        )
        self.data_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 必填欄位樣式
        self.required_style = NamedStyle(name='required')
        self.required_style.font = Font(bold=True, color='FF0000', size=11)
        self.required_style.fill = PatternFill(
            start_color='FFEEEE',
            end_color='FFEEEE',
            fill_type='solid'
        )
    
    def process_json_to_excel(self, json_data: str, output_path: Optional[str] = None) -> str:
        """
        將JSON資料轉換為Excel檔案
        
        Args:
            json_data: JSON格式的申請資料
            output_path: 輸出路徑（可選）
        
        Returns:
            產生的Excel檔案路徑
        """
        logger.info("開始處理JSON資料")
        
        # 解析JSON資料
        try:
            if isinstance(json_data, str):
                if os.path.isfile(json_data):
                    with open(json_data, 'r', encoding='utf-8') as f:
                        applications = json.load(f)
                else:
                    applications = json.loads(json_data)
            else:
                applications = json_data
        except Exception as e:
            logger.error(f"JSON解析失敗: {e}")
            raise
        
        logger.info(f"解析到 {len(applications)} 筆申請資料")
        
        # 按類別分組
        categorized = self._categorize_applications(applications)
        
        # 建立Excel檔案
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"SAP_Material_Import_{timestamp}.xlsx"
        
        # 建立工作簿
        wb = openpyxl.Workbook()
        
        # 註冊樣式
        if self.header_style.name not in wb.named_styles:
            wb.add_named_style(self.header_style)
        if self.data_style.name not in wb.named_styles:
            wb.add_named_style(self.data_style)
        if self.required_style.name not in wb.named_styles:
            wb.add_named_style(self.required_style)
        
        # 移除預設工作表
        wb.remove(wb.active)
        
        # 建立摘要工作表（如果啟用）
        if self.config['include_summary']:
            self._create_summary_sheet(wb, categorized, applications)
        
        # 為每個類別建立工作表
        for category_code, apps in categorized.items():
            if apps:
                sheet_name = self.category_mapping.get(category_code, 'Others')
                logger.info(f"建立工作表: {sheet_name} ({len(apps)} 筆資料)")
                ws = wb.create_sheet(sheet_name)
                self._write_category_sheet(ws, sheet_name, apps)
        
        # 儲存檔案
        try:
            wb.save(output_path)
            logger.info(f"✅ Excel檔案已成功產生: {output_path}")
        except Exception as e:
            logger.error(f"儲存檔案失敗: {e}")
            raise
        
        return output_path
    
    def _categorize_applications(self, applications: List[Dict]) -> Dict[str, List]:
        """按類別分組申請資料"""
        categorized = {}
        for app in applications:
            # 只處理已核准的申請
            if app.get('status') != 'APPROVED':
                continue
                
            category = app.get('mainCategory', 'O')
            if category not in categorized:
                categorized[category] = []
            categorized[category].append(app)
        return categorized
    
    def _create_summary_sheet(self, wb, categorized: Dict, applications: List):
        """建立摘要工作表"""
        ws = wb.create_sheet('Summary', 0)
        
        # 標題
        ws.merge_cells('A1:F1')
        ws['A1'] = '物料申請匯出摘要'
        ws['A1'].font = Font(bold=True, size=16, color='366092')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 基本資訊
        ws['A3'] = '匯出日期：'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws['A4'] = '總申請數：'
        ws['B4'] = len(applications)
        
        ws['A5'] = '已核准數：'
        approved = [a for a in applications if a.get('status') == 'APPROVED']
        ws['B5'] = len(approved)
        
        # 類別統計
        ws['A7'] = '類別統計'
        ws['A7'].font = Font(bold=True, size=12)
        
        row = 8
        ws[f'A{row}'] = '類別'
        ws[f'B{row}'] = '數量'
        ws[f'C{row}'] = '百分比'
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].style = 'header'
        
        row += 1
        total_approved = len(approved)
        
        for category_code, apps in categorized.items():
            category_name = self.category_mapping.get(category_code, 'Others')
            ws[f'A{row}'] = category_name
            ws[f'B{row}'] = len(apps)
            if total_approved > 0:
                ws[f'C{row}'] = f"{(len(apps) / total_approved * 100):.1f}%"
            else:
                ws[f'C{row}'] = "0%"
            row += 1
        
        # 調整欄寬
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 20
    
    def _write_category_sheet(self, ws, sheet_name: str, applications: List[Dict]):
        """寫入特定類別的工作表"""
        # 取得該類別的欄位
        columns = self.category_columns.get(sheet_name, self.category_columns['Others'])
        
        # 寫入標題
        for col_idx, column_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.style = 'header'
        
        # 寫入資料
        for row_idx, app in enumerate(applications, 2):
            self._write_application_row(ws, row_idx, app, columns)
        
        # 設定自動篩選
        if self.config['auto_filter']:
            ws.auto_filter.ref = ws.dimensions
        
        # 凍結窗格
        if self.config['freeze_panes']:
            ws.freeze_panes = self.config['freeze_panes']
        
        # 調整欄寬
        self._adjust_column_widths(ws, columns)
        
        # 加入資料驗證（如果啟用）
        if self.config['include_validation']:
            self._add_data_validation(ws, len(applications) + 1)
    
    def _write_application_row(self, ws, row_idx: int, app: Dict, columns: List[str]):
        """寫入單筆申請資料"""
        # 資料映射
        data_mapping = self._create_data_mapping(app)
        
        # 寫入每個欄位
        for col_idx, column_name in enumerate(columns, 1):
            value = data_mapping.get(column_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.style = 'data'
            
            # 特殊格式處理
            if '重量' in column_name or '長度' in column_name or '寬度' in column_name or '高度' in column_name:
                try:
                    if value:
                        cell.value = float(value)
                        cell.number_format = '#,##0.00'
                except:
                    pass
            elif '日期' in column_name:
                if value:
                    try:
                        date_obj = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        cell.value = date_obj.strftime(self.config['date_format'])
                    except:
                        pass
    
    def _create_data_mapping(self, app: Dict) -> Dict[str, Any]:
        """建立資料映射"""
        # 基本資料映射
        data_mapping = {
            '料號': app.get('itemCode', ''),
            '料件說明': app.get('itemNameCN', ''),
            '客戶說明': app.get('itemNameEN', ''),
            '產品大類': self.category_mapping.get(app.get('mainCategory', ''), ''),
            '產品中類': app.get('subCategory', ''),
            '產品小類': app.get('specCategory', ''),
            '料件基本材質': app.get('material', ''),
            '料件外型長': app.get('dimensions', {}).get('length', ''),
            '料件外型寬': app.get('dimensions', {}).get('width', ''),
            '料件外型高': app.get('dimensions', {}).get('height', ''),
            '料件外型重量': app.get('dimensions', {}).get('weight', ''),
            '料件表面處理': app.get('surfaceFinish', ''),
            'MOQ': app.get('moq', ''),
            '單位': app.get('unit', 'PCS'),
            '客戶參考號': app.get('customerRef', ''),
            '供應商編號': app.get('supplier', ''),
            '建立日期': app.get('submitDate', ''),
            '狀態': app.get('status', '')
        }
        
        # 包裝資料映射
        packaging = app.get('packaging', {})
        for key in ['個別產品包裝', '配件內容', '配件', '內盒', '外箱', '運輸與托盤要求', '裝櫃要求', 'Other']:
            # 相容性處理
            packaging_key = key
            if key == 'Other':
                packaging_key = '其他說明'
            
            field_data = packaging.get(packaging_key, packaging.get(key, {}))
            data_mapping[key] = self._format_packaging_field(field_data)
        
        # 特殊欄位處理（根據類別）
        main_category = app.get('mainCategory', '')
        if main_category == 'H':
            # Handle特殊欄位
            data_mapping['把手長度'] = app.get('dimensions', {}).get('length', '')
            data_mapping['孔距'] = app.get('handleHoleDistance', '')
        elif main_category == 'S':
            # Slide特殊欄位
            data_mapping['滑軌長度'] = app.get('dimensions', {}).get('length', '')
            data_mapping['滑軌載重'] = app.get('slideLoad', '')
            data_mapping['滑軌類型'] = app.get('slideType', '')
            data_mapping['鋼珠大小'] = app.get('ballSize', '')
        
        return data_mapping
    
    def _format_packaging_field(self, field_data: Any) -> str:
        """格式化包裝欄位資料"""
        if not field_data:
            return ''
        
        # 處理字串格式
        if isinstance(field_data, str):
            return field_data
        
        # 處理字典格式
        if isinstance(field_data, dict):
            options = field_data.get('options', [])
            description = field_data.get('description', '')
            
            result_parts = []
            if options:
                result_parts.append(f"[{', '.join(options)}]")
            if description:
                result_parts.append(description)
            
            return ' | '.join(result_parts) if result_parts else ''
        
        # 處理列表格式
        if isinstance(field_data, list):
            return ', '.join(str(item) for item in field_data)
        
        return str(field_data)
    
    def _adjust_column_widths(self, ws, columns: List[str]):
        """自動調整欄寬"""
        for col_idx, column_name in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            
            # 設定基本寬度
            if '說明' in column_name or '包裝' in column_name:
                width = 30
            elif '料號' in column_name or '編號' in column_name:
                width = 15
            elif any(keyword in column_name for keyword in ['長', '寬', '高', '重量', 'MOQ']):
                width = 10
            else:
                width = 12
            
            ws.column_dimensions[col_letter].width = width
    
    def _add_data_validation(self, ws, max_row: int):
        """加入資料驗證"""
        # 單位下拉選單
        unit_validation = DataValidation(
            type="list",
            formula1='"PCS,SET,PAIR,KG,M,BOX"',
            allow_blank=True
        )
        unit_validation.error = '請選擇有效的單位'
        unit_validation.errorTitle = '單位錯誤'
        
        # 找出單位欄位
        unit_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == '單位':
                unit_col = get_column_letter(col_idx)
                break
        
        if unit_col:
            unit_validation.add(f'{unit_col}2:{unit_col}{max_row}')
            ws.add_data_validation(unit_validation)
    
    def validate_excel_format(self, file_path: str) -> Dict[str, Any]:
        """
        驗證Excel檔案格式
        
        Args:
            file_path: Excel檔案路徑
        
        Returns:
            驗證結果字典
        """
        results = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'summary': {},
            'details': []
        }
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Summary':
                    continue
                    
                ws = wb[sheet_name]
                
                # 檢查欄位
                expected_columns = self.category_columns.get(sheet_name)
                if not expected_columns:
                    results['warnings'].append(f"工作表 {sheet_name} 非標準類別")
                    continue
                
                # 取得實際欄位
                actual_columns = []
                for cell in ws[1]:
                    if cell.value:
                        actual_columns.append(cell.value)
                
                # 檢查必要欄位
                required_columns = ['料號', '料件說明', '料件基本材質']
                missing_required = set(required_columns) - set(actual_columns)
                if missing_required:
                    results['errors'].append(f"工作表 {sheet_name} 缺少必要欄位: {missing_required}")
                    results['valid'] = False
                
                # 檢查其他欄位
                missing_columns = set(expected_columns) - set(actual_columns)
                if missing_columns:
                    results['warnings'].append(f"工作表 {sheet_name} 缺少欄位: {missing_columns}")
                
                # 統計資料
                row_count = 0
                for row in ws.iter_rows(min_row=2):
                    if any(cell.value for cell in row):
                        row_count += 1
                
                results['summary'][sheet_name] = row_count
                results['details'].append({
                    'sheet': sheet_name,
                    'rows': row_count,
                    'columns': len(actual_columns),
                    'missing_columns': list(missing_columns)
                })
            
            wb.close()
            
        except Exception as e:
            results['errors'].append(f"檔案讀取錯誤: {str(e)}")
            results['valid'] = False
        
        return results
    
    def merge_excel_files(self, file_paths: List[str], output_path: str) -> str:
        """
        合併多個Excel檔案
        
        Args:
            file_paths: Excel檔案路徑列表
            output_path: 輸出檔案路徑
        
        Returns:
            合併後的檔案路徑
        """
        logger.info(f"開始合併 {len(file_paths)} 個檔案")
        
        merged_data = {}
        
        for file_path in file_paths:
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                
                for sheet_name in wb.sheetnames:
                    if sheet_name == 'Summary':
                        continue
                    
                    if sheet_name not in merged_data:
                        merged_data[sheet_name] = []
                    
                    ws = wb[sheet_name]
                    
                    # 讀取標題
                    headers = []
                    for cell in ws[1]:
                        if cell.value:
                            headers.append(cell.value)
                    
                    # 讀取資料
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(row):
                            row_dict = dict(zip(headers, row))
                            merged_data[sheet_name].append(row_dict)
                
                wb.close()
                logger.info(f"成功讀取: {file_path}")
                
            except Exception as e:
                logger.error(f"讀取檔案失敗 {file_path}: {e}")
        
        # 建立新的工作簿
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 寫入合併的資料
        for sheet_name, data in merged_data.items():
            if data:
                ws = wb.create_sheet(sheet_name)
                
                # 寫入標題
                headers = list(data[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.style = 'header'
                
                # 寫入資料
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        value = row_data.get(header, '')
                        ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 儲存檔案
        wb.save(output_path)
        logger.info(f"✅ 合併完成: {output_path}")
        
        return output_path


def main():
    """主程式"""
    parser = argparse.ArgumentParser(
        description='物料編碼申請管理系統 V3.5 Excel處理器'
    )
    parser.add_argument(
        'action',
        choices=['convert', 'validate', 'merge'],
        help='執行動作'
    )
    parser.add_argument(
        '-i', '--input',
        help='輸入檔案路徑'
    )
    parser.add_argument(
        '-o', '--output',
        help='輸出檔案路徑'
    )
    parser.add_argument(
        '-c', '--config',
        help='配置檔案路徑'
    )
    parser.add_argument(
        '--files',
        nargs='+',
        help='要合併的檔案列表（用於merge動作）'
    )
    
    args = parser.parse_args()
    
    # 建立處理器
    processor = MaterialExcelProcessor(config_path=args.config)
    
    try:
        if args.action == 'convert':
            # 轉換JSON為Excel
            if not args.input:
                print("錯誤：請指定輸入檔案 (-i)")
                sys.exit(1)
            
            output_path = processor.process_json_to_excel(
                args.input,
                args.output
            )
            print(f"✅ 轉換完成: {output_path}")
            
        elif args.action == 'validate':
            # 驗證Excel格式
            if not args.input:
                print("錯誤：請指定要驗證的檔案 (-i)")
                sys.exit(1)
            
            results = processor.validate_excel_format(args.input)
            
            print("\n" + "="*60)
            print("驗證結果")
            print("="*60)
            print(f"✅ 檔案有效: {results['valid']}")
            print(f"📊 資料統計: {results['summary']}")
            
            if results['errors']:
                print(f"\n❌ 錯誤:")
                for error in results['errors']:
                    print(f"  - {error}")
            
            if results['warnings']:
                print(f"\n⚠️ 警告:")
                for warning in results['warnings']:
                    print(f"  - {warning}")
            
            print("\n詳細資訊:")
            for detail in results['details']:
                print(f"  {detail['sheet']}: {detail['rows']} 筆資料, {detail['columns']} 個欄位")
            
        elif args.action == 'merge':
            # 合併多個Excel檔案
            if not args.files or len(args.files) < 2:
                print("錯誤：請指定至少2個要合併的檔案 (--files)")
                sys.exit(1)
            
            if not args.output:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                args.output = f"Merged_{timestamp}.xlsx"
            
            output_path = processor.merge_excel_files(
                args.files,
                args.output
            )
            print(f"✅ 合併完成: {output_path}")
        
    except Exception as e:
        print(f"❌ 錯誤: {e}")
        logger.exception("處理失敗")
        sys.exit(1)


if __name__ == "__main__":
    # 測試模式
    if len(sys.argv) == 1:
        print("="*60)
        print("物料編碼申請管理系統 V3.5 Excel處理器 - 測試模式")
        print("="*60)
        
        # 建立測試資料
        test_data = [
            {
                'id': '1700000001',
                'submitDate': datetime.now().isoformat(),
                'status': 'APPROVED',
                'itemCode': 'H01.C.00001',
                'mainCategory': 'H',
                'subCategory': '01',
                'specCategory': 'C',
                'itemNameCN': '鍍鉻把手 160mm',
                'itemNameEN': 'Chrome Handle 160mm',
                'customerRef': 'CUST-001',
                'supplier': 'SUP001',
                'material': 'Zinc Alloy',
                'surfaceFinish': 'Chrome Plated',
                'dimensions': {
                    'length': 160,
                    'width': 25,
                    'height': 35,
                    'weight': 120
                },
                'moq': 500,
                'unit': 'PCS',
                'packaging': {
                    '產品包裝': {
                        'options': ['塑膠袋', 'PE/PP材質', '產品標籤'],
                        'description': '1PC/塑膠袋，印刷回收標誌04 PE-LD'
                    },
                    '配件內容': {
                        'options': ['螺絲'],
                        'description': '附M4x25mm螺絲2顆'
                    },
                    '配件': {
                        'options': [],
                        'description': ''
                    },
                    '內盒': {
                        'options': ['印製ITEM NO.', '印製數量'],
                        'description': '內盒印製產品編號及數量'
                    },
                    '外箱': {
                        'options': ['瓦楞紙箱', '側嘜'],
                        'description': '5層瓦楞紙箱，側嘜印製客戶編號'
                    },
                    '運輸與托盤要求': {
                        'options': ['托盤/Pallet', 'EUDR文件'],
                        'description': '歐規托盤，出貨提供EUDR文件'
                    },
                    '裝櫃要求': {
                        'options': ['40呎櫃'],
                        'description': '標準40呎貨櫃'
                    },
                    '其他說明': {
                        'options': ['FSC認證'],
                        'description': '供應商需具備FSC認證'
                    }
                }
            }
        ]
        
        # 測試轉換
        processor = MaterialExcelProcessor()
        output_file = processor.process_json_to_excel(test_data)
        
        # 測試驗證
        validation = processor.validate_excel_format(output_file)
        
        print("\n測試結果:")
        print(f"✅ 檔案產生: {output_file}")
        print(f"✅ 格式驗證: {'通過' if validation['valid'] else '失敗'}")
        print(f"📊 資料統計: {validation['summary']}")
        
        print("\n使用說明:")
        print("python excel_processor.py convert -i input.json -o output.xlsx")
        print("python excel_processor.py validate -i file.xlsx")
        print("python excel_processor.py merge --files file1.xlsx file2.xlsx -o merged.xlsx")
    else:
        main()
